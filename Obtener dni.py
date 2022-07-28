import mysql.connector
import openpyxl
wb = openpyxl.Workbook()
hoja = wb.active
hoja.title = "Datos"

def listar_fechas(dni):
    
    sql = f"SELECT * FROM personas_sistemas WHERE dni = {dni}";
    cursor.execute(sql)
    result = cursor.fetchall()
    lista_fechas = []
    result.sort(key = lambda x: x[10], reverse=True)
    print(result)
    for fecha in result:
        lista_fechas.append(fecha[10])

    print('')
    print(f'Para el DNI {dni}')
    print('Las fechas ordenadas de mas actual a mas antigua son: ')
    print('')
    lista_fechas.sort(reverse=True)
    print(lista_fechas)
    print('')
    
   # print(f"La fecha mas reciente es: {lista_fechas[10]}")
    print('')
    print('Modificando Excel...')

    try:
        hoja.cell(row,1).value = dni
        print("Dni OK")
        hoja.cell(row,2).value = result[0][1]
        print("Apellido OK")
        hoja.cell(row,3).value = result[0][2]
        print("Nombre OK")
        hoja.cell(row,4).value = result[0][6]
        print("Celular OK")
        hoja.cell(row,5).value = result[0][7]
        print("Localidad OK")
        hoja.cell(row,6).value = result[0][8]
        print("Barrio OK")
        hoja.cell(row,7).value = lista_fechas[0]
        print("Alta OK")
    
    except:
        print("Algo falló")
    
    wb.save("datos_sql_obtenidos.xlsx")

hoja["A1"] = 'DNI'
hoja["B1"] = 'Apellido'
hoja["C1"] = 'Nombre'
hoja["D1"] = 'Celular'
hoja["E1"] = 'Localidad'
hoja["F1"] = 'Barrio'
hoja["G1"] = 'Alta'

row = 2

db = { 
    'host' : 'EJEMPLO',
        'database' : 'EJEMPLO',
        'user' : 'EJEMPLO',
        'password' : 'EJEMPLO',
        'port' : 'EJEMPLO',
    }

conexion = mysql.connector.connect(**db)

cursor = conexion.cursor()

sql = "SELECT distinct dni FROM personas_sistemas where dni !='' and dni is not null and celular != ' ' and alta != ' ' "
cursor.execute(sql)
result = cursor.fetchall()
lista_dni = []

for dni in result:
    lista_dni.append(dni)

lista_dni.sort()

for dni in lista_dni:
    listar_fechas(dni[0])
    row += 1

print("Fin de la ejecucion")
wb.save("datos_sql_obtenidos.xlsx")